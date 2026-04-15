#!/usr/bin/env python3
"""Convert Pastor Luis's 2025 consolidated xlsx into 12 relatorio_mensal_2025-MM.md files.

This is a one-shot importer — the 2026 flow uses /tesouraria-fechar, not this script.
Run once, commit the resulting 12 folders, then discard.
"""
from __future__ import annotations

import argparse
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

XLSX_DEFAULT = Path.home() / "Data/Documents/Finance/2025/Planilha de Acompanhamento Financeiro - 2025.xlsx"
FINANCE_ROOT = Path.home() / "Data/Documents/Finance/Tesouraria-PIB-Confins"

# Monthly sheets are at workbook indices 2..13 (1-based, skipping global + dizimos sheets).
# Using index rather than name because sheets 11-13 have "Mensaa" typo.
MONTHLY_SHEET_INDEX = list(range(2, 14))  # 12 sheets

MONTH_NAMES = {
    1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
    5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
    9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro",
}

MONTH_SLUG = {
    1: "janeiro", 2: "fevereiro", 3: "marco", 4: "abril",
    5: "maio", 6: "junho", 7: "julho", 8: "agosto",
    9: "setembro", 10: "outubro", 11: "novembro", 12: "dezembro",
}

# Row coordinates inside each monthly sheet (column G = 7 holds all values)
ROW_MEMBROS = (4, 54)     # inclusive
ROW_TOTAL_MEMBROS = 58
ROW_NAO_MEMBROS_START = 60
ROW_TOTAL_NAO_MEMBROS = 98
ROW_OFERTAS = {
    "anonimas": 100, "missoes": 101, "recantoVida": 102,
    "beneficencia": 103, "outros": 104,
}
ROW_TOTAL_OFERTAS = 105
ROW_SALDO_ANTERIOR = 107  # col C
ROW_TOTAL_ENTRADAS_MES = 107  # col G
ROW_SOMATORIO_ENTRADAS = 109  # col G
ROW_SAIDAS = {
    "planoCooperativo": 115, "zeladoria": 116, "sustentoMinisterial": 117,
    "missoes": 118, "copasa": 119, "cemig": 120,
    "manutencao": 121, "despesasEventuais": 122, "outros": 123,
}
ROW_TOTAL_SAIDAS = 128
ROW_RESUMO_SALDO_MES = 132
ROW_RESUMO_NOVO_SALDO = 135
COL_VALUE = 7  # G
COL_LABEL = 2  # B (used for membros/nao-membros sections)
COL_LABEL_A = 1  # A (used for saidas section — labels are in col A there)
COL_SALDO_ANT = 3  # C


def cell_num(ws, row: int, col: int) -> float | None:
    v = ws.cell(row, col).value
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def extract_membros(ws) -> tuple[list[tuple[str, float | None]], float]:
    """Return (list of (name, value|None), printed total)."""
    result = []
    for row in range(ROW_MEMBROS[0], ROW_MEMBROS[1] + 1):
        name = ws.cell(row, COL_LABEL).value
        if not name:
            continue
        value = cell_num(ws, row, COL_VALUE)
        result.append((str(name).strip(), value))
    total = cell_num(ws, ROW_TOTAL_MEMBROS, COL_VALUE) or 0.0
    return result, total


def extract_nao_membros(ws) -> tuple[list[tuple[str, float | None]], float]:
    """Non-members section runs from row 60 until the TOTAL row (98).
    Rows are numbered #1..#N in col A; some rows (e.g. 94-97) have compound labels.
    Skip rows without a name in col B."""
    result = []
    for row in range(ROW_NAO_MEMBROS_START, ROW_TOTAL_NAO_MEMBROS):
        name = ws.cell(row, COL_LABEL).value
        if not name or str(name).strip().upper() == "TOTAL":
            continue
        value = cell_num(ws, row, COL_VALUE)
        result.append((str(name).strip(), value))
    total = cell_num(ws, ROW_TOTAL_NAO_MEMBROS, COL_VALUE) or 0.0
    return result, total


def extract_saidas(ws) -> dict:
    """Named categories are at fixed rows. Rows 123+ may contain 'Outros' label plus
    custom labels (e.g. '1/3 Férias do Obreiro') — collect all non-empty rows between
    the named 'Outros' (123) and TOTAL (128) as `outros` list.
    Note: saidas labels are in col A (not col B like membros/nao-membros)."""
    result = {}
    for key, row in ROW_SAIDAS.items():
        if key == "outros":
            continue
        v = cell_num(ws, row, COL_VALUE)
        # scalar fields that default to 0 when absent
        if key in ("planoCooperativo", "zeladoria", "sustentoMinisterial", "copasa"):
            result[key] = v if v is not None else 0.0
        else:
            result[key] = v if v is not None else 0.0
    # "Outros" and any custom rows between 123..127
    outros = []
    for row in range(ROW_SAIDAS["outros"], ROW_TOTAL_SAIDAS):
        label = ws.cell(row, COL_LABEL_A).value
        v = cell_num(ws, row, COL_VALUE)
        if not label or v is None or v == 0:
            continue
        outros.append({"descricao": str(label).strip(), "valor": round(v, 2)})
    result["outros"] = outros
    result["total"] = cell_num(ws, ROW_TOTAL_SAIDAS, COL_VALUE) or 0.0
    return result


def extract_resumo(ws) -> dict:
    return {
        "saldoMes": cell_num(ws, ROW_RESUMO_SALDO_MES, COL_VALUE),
        "novoSaldo": cell_num(ws, ROW_RESUMO_NOVO_SALDO, COL_VALUE),
    }


def extract_ofertas(ws) -> tuple[dict, float]:
    ofertas = {k: cell_num(ws, row, COL_VALUE) for k, row in ROW_OFERTAS.items()}
    # Normalize None to 0.0 for ofertas (they have explicit 0 in xlsx, but be safe)
    ofertas = {k: (v if v is not None else 0.0) for k, v in ofertas.items()}
    total = cell_num(ws, ROW_TOTAL_OFERTAS, COL_VALUE) or 0.0
    return ofertas, total


def extract_entrada_geral(ws) -> dict:
    return {
        "saldoAnterior": cell_num(ws, ROW_SALDO_ANTERIOR, COL_SALDO_ANT),
        "totalMes": cell_num(ws, ROW_TOTAL_ENTRADAS_MES, COL_VALUE),
        "somatorio": cell_num(ws, ROW_SOMATORIO_ENTRADAS, COL_VALUE),
    }


@dataclass
class MesData:
    month: int
    membros: list[tuple[str, float | None]]
    total_membros: float
    nao_membros: list[tuple[str, float | None]]
    total_nao_membros: float
    ofertas: dict
    total_ofertas: float
    entrada_geral: dict
    saidas: dict
    resumo: dict


def extract_month(ws, month: int) -> MesData:
    membros, total_m = extract_membros(ws)
    nao_m, total_nm = extract_nao_membros(ws)
    of, total_of = extract_ofertas(ws)
    return MesData(
        month=month,
        membros=membros, total_membros=total_m,
        nao_membros=nao_m, total_nao_membros=total_nm,
        ofertas=of, total_ofertas=total_of,
        entrada_geral=extract_entrada_geral(ws),
        saidas=extract_saidas(ws),
        resumo=extract_resumo(ws),
    )


def fmt_brl(v: float | None) -> str:
    """1234.56 -> '1.234,56'; None -> '—'."""
    if v is None:
        return "—"
    negative = v < 0
    s = f"{abs(v):,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")
    return f"-{s}" if negative else s


def _pessoa_row(i: int, nome: str, valor: float | None) -> str:
    return f"| {i} | {nome} | {fmt_brl(valor)} |"


def render_md(d: MesData, month: int) -> str:
    mes_nome = MONTH_NAMES[month]
    lines = [
        f"# Relatório Financeiro — Congregação em Confins — {mes_nome}/2025",
        "",
        "> **Gerado a partir da Planilha 2025 consolidada de Pr. Luis em 2026-04-15.**",
        "> Totais derivados da planilha; comprovantes individuais não disponíveis neste histórico.",
        "",
        "---",
        "",
        "## ENTRADAS — Membros",
        "",
        "| # | Nome | Valor (R$) |",
        "|---|------|-----------:|",
    ]
    for i, (nome, valor) in enumerate(d.membros, start=1):
        lines.append(_pessoa_row(i, nome, valor))
    lines.append(f"| | **TOTAL MEMBROS** | **R$ {fmt_brl(d.total_membros)}** |")
    lines += ["", "## ENTRADAS — Não Membros", "",
              "| # | Nome | Valor (R$) |",
              "|---|------|-----------:|"]
    for i, (nome, valor) in enumerate(d.nao_membros, start=1):
        lines.append(_pessoa_row(i, nome, valor))
    lines.append(f"| | **TOTAL NÃO MEMBROS** | **R$ {fmt_brl(d.total_nao_membros)}** |")
    lines += ["", "## OFERTAS", "",
              "| Tipo | Valor (R$) |", "|------|-----------:|",
              f"| Anônimas | {fmt_brl(d.ofertas['anonimas'])} |",
              f"| Missões | {fmt_brl(d.ofertas['missoes'])} |",
              f"| Recanto Vida | {fmt_brl(d.ofertas['recantoVida'])} |",
              f"| Beneficência | {fmt_brl(d.ofertas['beneficencia'])} |",
              f"| Outros | {fmt_brl(d.ofertas['outros'])} |",
              f"| **TOTAL OFERTAS** | **R$ {fmt_brl(d.total_ofertas)}** |"]
    eg = d.entrada_geral
    lines += ["", "## ENTRADA GERAL", "",
              "| | |", "|---|---:|",
              f"| Saldo Recorrente do Mês Anterior | R$ {fmt_brl(eg['saldoAnterior'])} |",
              f"| Total de Entradas Deste Mês | R$ {fmt_brl(eg['totalMes'])} |",
              f"| **Somatório Geral de Entradas** | **R$ {fmt_brl(eg['somatorio'])}** |"]
    s = d.saidas
    lines += ["", "## SAÍDAS", "",
              "| Especificação | Valor (R$) |", "|---|---:|",
              f"| Plano Cooperativo | {fmt_brl(s['planoCooperativo'])} |",
              f"| Zeladoria (1/3 do salário vigente) | {fmt_brl(s['zeladoria'])} |",
              f"| Sustento Ministerial | {fmt_brl(s['sustentoMinisterial'])} |",
              f"| Missões | {fmt_brl(s['missoes'])} |",
              f"| COPASA | {fmt_brl(s['copasa'])} |",
              f"| CEMIG | {fmt_brl(s['cemig'])} |",
              f"| Manutenção e Limpeza | {fmt_brl(s['manutencao'])} |",
              f"| Despesas Eventuais | {fmt_brl(s['despesasEventuais'])} |"]
    for o in s["outros"]:
        lines.append(f"| {o['descricao']} | {fmt_brl(o['valor'])} |")
    lines.append(f"| **TOTAL SAÍDAS** | **R$ {fmt_brl(s['total'])}** |")
    r = d.resumo
    lines += ["", "## RESUMO DE CAIXA", "",
              "| | |", "|---|---:|",
              f"| Total de Entrada do Mês | R$ {fmt_brl(eg['totalMes'])} |",
              f"| Total de Saídas do Mês | R$ {fmt_brl(s['total'])} |",
              f"| **Saldo do Mês** | **R$ {fmt_brl(r['saldoMes'])}** |",
              f"| Saldo Recorrente do Mês Anterior | R$ {fmt_brl(eg['saldoAnterior'])} |",
              f"| **Novo Saldo Para o Próximo Mês** | **R$ {fmt_brl(r['novoSaldo'])}** |",
              "", "## Assinaturas", "",
              "- Presidente: Luis Gustavo Amaral Muritiba ___________________________",
              "- 1º Tesoureiro: ___________________________________________________",
              "- 2º Tesoureiro: ___________________________________________________"]
    return "\n".join(lines) + "\n"


def write_all(xlsx: Path, out_root: Path, dry_run: bool = False) -> None:
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    anual_ent = 0.0
    anual_sai = 0.0
    for idx in range(12):
        month = idx + 1
        ws = wb.worksheets[idx + 2]
        data = extract_month(ws, month)
        md = render_md(data, month)
        folder = out_root / f"2025-{month:02d}-{MONTH_SLUG[month]}"
        target = folder / f"relatorio_mensal_2025-{month:02d}.md"
        anual_ent += data.entrada_geral["totalMes"] or 0
        anual_sai += data.saidas["total"]
        if dry_run:
            print(f"would write {target}")
            continue
        folder.mkdir(parents=True, exist_ok=True)
        target.write_text(md, encoding="utf-8")
        print(f"✓ wrote {target}")
    print(f"\nAnnual totals: entradas={anual_ent:.2f}  saídas={anual_sai:.2f}")
    assert abs(anual_ent - 104539.56) < 0.01, f"entradas mismatch: {anual_ent}"
    assert abs(anual_sai - 92137.54) < 0.01, f"saídas mismatch: {anual_sai}"
    print("✓ annual totals reconcile with global sheet")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=XLSX_DEFAULT)
    ap.add_argument("--out", type=Path, default=FINANCE_ROOT)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()
    write_all(args.xlsx, args.out, args.dry_run)
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())
